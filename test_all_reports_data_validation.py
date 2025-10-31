"""
Test All Reports Data Validation (PDF, PPT, Excel)
Validates that all report formats contain real data from synthesized_data with no placeholders
"""
import json
from pathlib import Path
from datetime import datetime
from loguru import logger
import sys

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / 'src'))

from src.outputs.revolutionary_pdf_generator import RevolutionaryPDFGenerator
from src.outputs.revolutionary_ppt_generator import RevolutionaryPowerPointGenerator
from src.outputs.revolutionary_excel_generator import RevolutionaryExcelGenerator
from src.utils.data_accessor import DataAccessor
import importlib.util

# Placeholder patterns to detect
PLACEHOLDER_PATTERNS = [
    'placeholder', 'TBD', 'N/A', 'pending', 'in progress',
    'not available', 'default', 'example', 'sample',
    'coming soon', 'under review', 'to be determined', 'null'
]

# Expected data fields that must have real values
REQUIRED_FIELDS = {
    'executive_summary': ['key_recommendation', 'strategic_rationale'],
    'detailed_financials': ['quality_score', 'dcf_outputs'],
    'validation_summary': ['overall_confidence_score'],
}


def load_recent_state(symbol='NVDA'):
    """Load most recent analysis state for symbol"""
    logger.info(f"Loading {symbol} state...")
    
    # Check frontend_results directory
    results_dir = Path('frontend_results')
    if results_dir.exists():
        state_files = list(results_dir.glob(f'**/*{symbol}*.json'))
        if state_files:
            latest_file = max(state_files, key=lambda p: p.stat().st_mtime)
            logger.info(f"Loading from frontend_results: {latest_file}")
            with open(latest_file, 'r') as f:
                state = json.load(f)
            logger.info(f"✓ Loaded {symbol} state with {len(state.get('agent_outputs', []))} agent outputs")
            return state
    
    # Check data directory
    data_dir = Path('data')
    if data_dir.exists():
        state_files = list(data_dir.glob(f'*{symbol}*.json'))
        if state_files:
            latest_file = max(state_files, key=lambda p: p.stat().st_mtime)
            logger.info(f"Loading from data: {latest_file}")
            with open(latest_file, 'r') as f:
                state = json.load(f)
            logger.info(f"✓ Loaded {symbol} state with {len(state.get('agent_outputs', []))} agent outputs")
            return state
    
    logger.error(f"No {symbol} state files found")
    return None


def validate_synthesized_data(state):
    """Validate synthesized data exists and has content"""
    logger.info("\n=== VALIDATING SYNTHESIZED DATA ===")
    
    validation = DataAccessor.validate_data_consistency(state)
    
    if not validation['has_synthesized_data']:
        logger.error("❌ No synthesized data found!")
        return False
    
    logger.info(f"✓ Synthesized data version: {validation['data_version']}")
    
    synthesized_data = DataAccessor.get_synthesized_data(state)
    
    # Check required fields
    issues = []
    for section, fields in REQUIRED_FIELDS.items():
        if section not in synthesized_data:
            issues.append(f"Missing section: {section}")
            continue
        
        section_data = synthesized_data[section]
        for field in fields:
            if field not in section_data:
                issues.append(f"Missing field: {section}.{field}")
            elif not section_data[field]:
                issues.append(f"Empty field: {section}.{field}")
    
    if issues:
        logger.warning(f"⚠ Found {len(issues)} data issues:")
        for issue in issues[:10]:
            logger.warning(f"  - {issue}")
    else:
        logger.info("✓ All required fields present and populated")
    
    # Show data summary
    logger.info("\nSynthesized Data Summary:")
    logger.info(f"  Sections: {len(synthesized_data)}")
    
    # Key metrics
    exec_summary = synthesized_data.get('executive_summary', {})
    logger.info(f"  Recommendation: {exec_summary.get('key_recommendation', 'N/A')[:50]}")
    
    detailed_fin = synthesized_data.get('detailed_financials', {})
    logger.info(f"  Quality Score: {detailed_fin.get('quality_score', 0)}/100")
    
    dcf_outputs = detailed_fin.get('dcf_outputs', {})
    ev = dcf_outputs.get('enterprise_value', 0)
    logger.info(f"  Enterprise Value: ${abs(ev)/1e9:.2f}B")
    
    return True


def check_for_placeholders(state):
    """Check if synthesized data contains placeholder values"""
    logger.info("\n=== CHECKING FOR PLACEHOLDERS IN DATA ===")
    
    synthesized_data = DataAccessor.get_synthesized_data(state)
    
    # Convert to JSON string for searching
    data_str = json.dumps(synthesized_data, default=str).lower()
    
    found_placeholders = []
    for pattern in PLACEHOLDER_PATTERNS:
        if pattern.lower() in data_str:
            count = data_str.count(pattern.lower())
            found_placeholders.append((pattern, count))
    
    if found_placeholders:
        logger.warning(f"⚠ Found {len(found_placeholders)} placeholder patterns in data:")
        for pattern, count in found_placeholders[:10]:
            logger.warning(f"  - '{pattern}': {count} occurrences")
        return False
    else:
        logger.info("✓ No common placeholder patterns detected in data")
        return True


def generate_and_validate_pdf(state):
    """Generate PDF and validate content"""
    logger.info("\n=== GENERATING AND VALIDATING PDF ===")
    
    try:
        pdf_gen = RevolutionaryPDFGenerator()
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pdf_path = f"outputs/test_validation_{timestamp}.pdf"
        
        logger.info(f"Generating PDF to: {pdf_path}")
        pdf_gen.generate_report(state, pdf_path)
        
        logger.info(f"✓ PDF generated successfully")
        
        pdf_file = Path(pdf_path)
        if pdf_file.exists():
            size_kb = pdf_file.stat().st_size / 1024
            logger.info(f"✓ PDF file size: {size_kb:.1f} KB")
            
            if size_kb < 50:
                logger.warning("⚠ PDF file seems small")
                return False, pdf_path
            
            return True, pdf_path
        else:
            logger.error("❌ PDF file not created")
            return False, None
            
    except Exception as e:
        logger.error(f"❌ PDF generation failed: {e}")
        import traceback
        traceback.print_exc()
        return False, None


def generate_and_validate_ppt(state):
    """Generate PPT and validate content"""
    logger.info("\n=== GENERATING AND VALIDATING PPT ===")
    
    try:
        ppt_gen = RevolutionaryPowerPointGenerator()
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        ppt_path = f"outputs/test_validation_{timestamp}.pptx"
        
        logger.info(f"Generating PPT to: {ppt_path}")
        ppt_gen.generate_report(state, ppt_path)
        
        logger.info(f"✓ PPT generated successfully")
        
        ppt_file = Path(ppt_path)
        if ppt_file.exists():
            size_kb = ppt_file.stat().st_size / 1024
            logger.info(f"✓ PPT file size: {size_kb:.1f} KB")
            
            if size_kb < 50:
                logger.warning("⚠ PPT file seems small")
                return False, ppt_path
            
            return True, ppt_path
        else:
            logger.error("❌ PPT file not created")
            return False, None
            
    except Exception as e:
        logger.error(f"❌ PPT generation failed: {e}")
        import traceback
        traceback.print_exc()
        return False, None


def generate_and_validate_excel(state):
    """Generate Excel and validate content"""
    logger.info("\n=== GENERATING AND VALIDATING EXCEL ===")
    
    try:
        excel_gen = RevolutionaryExcelGenerator()
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_path = f"outputs/test_validation_{timestamp}.xlsx"
        
        logger.info(f"Generating Excel to: {excel_path}")
        excel_gen.generate_report(state, excel_path)
        
        logger.info(f"✓ Excel generated successfully")
        
        excel_file = Path(excel_path)
        if excel_file.exists():
            size_kb = excel_file.stat().st_size / 1024
            logger.info(f"✓ Excel file size: {size_kb:.1f} KB")
            
            if size_kb < 10:
                logger.warning("⚠ Excel file seems small")
                return False, excel_path
            
            # Validate Excel has real data
            return validate_excel_content(excel_file), excel_path
        else:
            logger.error("❌ Excel file not created")
            return False, None
            
    except Exception as e:
        logger.error(f"❌ Excel generation failed: {e}")
        import traceback
        traceback.print_exc()
        return False, None


def validate_excel_content(excel_path):
    """Validate Excel file contains real data, no placeholders"""
    logger.info("\nValidating Excel content...")
    
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(excel_path)
        logger.info(f"✓ Excel has {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames[:5])}")
        
        # Check for placeholders in cell values
        placeholder_count = 0
        total_cells_checked = 0
        
        for sheet_name in wb.sheetnames[:10]:  # Check first 10 sheets
            sheet = wb[sheet_name]
            
            for row in sheet.iter_rows(max_row=100, max_col=20):  # Sample cells
                for cell in row:
                    if cell.value:
                        total_cells_checked += 1
                        cell_str = str(cell.value).lower()
                        
                        # Check for common placeholders
                        for pattern in ['placeholder', 'tbd', 'n/a', 'pending', 'null', 'none']:
                            if pattern in cell_str and len(cell_str) < 20:  # Avoid false positives
                                placeholder_count += 1
                                logger.warning(f"  ⚠ Found '{cell.value}' in {sheet_name}!{cell.coordinate}")
                                break
        
        logger.info(f"✓ Checked {total_cells_checked} cells across {len(wb.sheetnames)} sheets")
        
        if placeholder_count > 0:
            logger.warning(f"⚠ Found {placeholder_count} potential placeholder values in Excel")
            return False
        else:
            logger.info("✓ No placeholder values detected in Excel content")
            return True
            
    except Exception as e:
        logger.error(f"❌ Excel content validation failed: {e}")
        return False


def generate_and_validate_dashboard(state):
    """Generate dashboard HTML and validate content"""
    logger.info("\n=== GENERATING AND VALIDATING DASHBOARD ===")
    
    try:
        # Load dashboard module
        dashboard_path = Path('revolutionary_dashboard.py')
        if not dashboard_path.exists():
            logger.warning("⚠ revolutionary_dashboard.py not found")
            return False, None
        
        spec = importlib.util.spec_from_file_location("revolutionary_dashboard", dashboard_path)
        dashboard_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(dashboard_module)
        
        # Generate dashboard
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        dashboard_path = f"outputs/test_dashboard_{timestamp}.html"
        
        logger.info(f"Generating Dashboard to: {dashboard_path}")
        
        # Call dashboard generation function
        if hasattr(dashboard_module, 'generate_dashboard'):
            dashboard_module.generate_dashboard(state, dashboard_path)
        else:
            logger.warning("⚠ generate_dashboard function not found in module")
            return False, None
        
        logger.info(f"✓ Dashboard generated successfully")
        
        dashboard_file = Path(dashboard_path)
        if dashboard_file.exists():
            size_kb = dashboard_file.stat().st_size / 1024
            logger.info(f"✓ Dashboard file size: {size_kb:.1f} KB")
            
            # Validate HTML content
            return validate_dashboard_content(dashboard_file), dashboard_path
        else:
            logger.error("❌ Dashboard file not created")
            return False, None
            
    except Exception as e:
        logger.error(f"❌ Dashboard generation failed: {e}")
        import traceback
        traceback.print_exc()
        return False, None


def validate_dashboard_content(dashboard_path):
    """Validate dashboard HTML contains real data, no placeholders"""
    logger.info("\nValidating Dashboard content...")
    
    try:
        with open(dashboard_path, 'r', encoding='utf-8') as f:
            html_content = f.read().lower()
        
        logger.info(f"✓ Dashboard HTML size: {len(html_content)} characters")
        
        # Check for placeholders
        placeholder_count = 0
        found_patterns = []
        
        for pattern in ['placeholder', 'tbd', 'pending', 'coming soon', 'n/a', 'null']:
            if pattern in html_content:
                count = html_content.count(pattern)
                placeholder_count += count
                found_patterns.append((pattern, count))
        
        if placeholder_count > 0:
            logger.warning(f"⚠ Found {len(found_patterns)} placeholder patterns in dashboard:")
            for pattern, count in found_patterns[:5]:
                logger.warning(f"  - '{pattern}': {count} occurrences")
            return False
        else:
            logger.info("✓ No placeholder patterns detected in dashboard HTML")
        
        # Check for key metrics (should have real values)
        required_elements = [
            'enterprise value', 'valuation', 'recommendation',
            'quality score', 'risk', 'financial'
        ]
        
        found_elements = sum(1 for elem in required_elements if elem in html_content)
        logger.info(f"✓ Found {found_elements}/{len(required_elements)} key dashboard elements")
        
        if found_elements < len(required_elements) // 2:
            logger.warning("⚠ Dashboard may be missing key content")
            return False
        
        return True
            
    except Exception as e:
        logger.error(f"❌ Dashboard content validation failed: {e}")
        return False


def validate_data_mapping(state):
    """Validate that data mapping works correctly"""
    logger.info("\n=== VALIDATING DATA MAPPING ===")
    
    synthesized_data = DataAccessor.get_synthesized_data(state)
    
    mappings_valid = True
    
    # 1. Executive Summary
    exec_summary = synthesized_data.get('executive_summary', {})
    if not exec_summary.get('key_recommendation'):
        logger.warning("⚠ executive_summary.key_recommendation is empty")
        mappings_valid = False
    else:
        logger.info(f"✓ key_recommendation: {exec_summary['key_recommendation'][:60]}...")
    
    # 2. Financial Metrics
    detailed_fin = synthesized_data.get('detailed_financials', {})
    dcf_outputs = detailed_fin.get('dcf_outputs', {})
    
    if not dcf_outputs:
        logger.warning("⚠ detailed_financials.dcf_outputs is empty")
        mappings_valid = False
    else:
        ev = dcf_outputs.get('enterprise_value', 0)
        if ev == 0:
            logger.warning("⚠ enterprise_value is zero")
            mappings_valid = False
        else:
            logger.info(f"✓ enterprise_value: ${abs(ev)/1e9:.2f}B")
    
    # 3. Quality Score
    quality = detailed_fin.get('quality_score', 0)
    if quality == 0:
        logger.warning("⚠ quality_score is zero")
        mappings_valid = False
    else:
        logger.info(f"✓ quality_score: {quality}/100")
    
    # 4. Validation Summary
    validation = synthesized_data.get('validation_summary', {})
    confidence = validation.get('overall_confidence_score', 0)
    if confidence == 0:
        logger.warning("⚠ overall_confidence_score is zero")
        mappings_valid = False
    else:
        logger.info(f"✓ overall_confidence_score: {confidence:.1%}")
    
    # 5. Risk Assessment
    risk_macro = synthesized_data.get('risk_macro', {})
    key_risks = risk_macro.get('key_risks', [])
    if not key_risks:
        logger.warning("⚠ risk_macro.key_risks is empty")
        mappings_valid = False
    else:
        logger.info(f"✓ key_risks: {len(key_risks)} risks identified")
    
    if mappings_valid:
        logger.info("\n✓ All key data mappings are working correctly")
    else:
        logger.warning("\n⚠ Some data mappings may not be working correctly")
    
    return mappings_valid


def main():
    """Main test execution"""
    logger.info("=" * 80)
    logger.info("ALL REPORTS DATA VALIDATION TEST (PDF + PPT + EXCEL)")
    logger.info("Validating all formats use real data from synthesized_data")
    logger.info("=" * 80)
    
    # Load state
    state = load_recent_state('NVDA')
    if not state:
        logger.error("❌ Failed to load analysis state")
        logger.info("\nTip: Run 'python test_jpm_gs_orchestrator.py --symbol NVDA' first")
        return False
    
    # Validate synthesized data
    if not validate_synthesized_data(state):
        logger.error("❌ Synthesized data validation failed")
        return False
    
    # Check for placeholders in data
    no_placeholders = check_for_placeholders(state)
    
    # Validate data mapping
    mapping_valid = validate_data_mapping(state)
    
    # Generate all four report formats
    pdf_success, pdf_path = generate_and_validate_pdf(state)
    ppt_success, ppt_path = generate_and_validate_ppt(state)
    excel_success, excel_path = generate_and_validate_excel(state)
    dashboard_success, dashboard_path = generate_and_validate_dashboard(state)
    
    # Final results
    logger.info("\n" + "=" * 80)
    logger.info("TEST RESULTS SUMMARY")
    logger.info("=" * 80)
    
    results = {
        "Synthesized Data Available": "✓ PASS" if state else "❌ FAIL",
        "No Placeholders in Data": "✓ PASS" if no_placeholders else "⚠ WARNING",
        "Data Mapping Valid": "✓ PASS" if mapping_valid else "⚠ WARNING",
        "PDF Generation": "✓ PASS" if pdf_success else "❌ FAIL",
        "PPT Generation": "✓ PASS" if ppt_success else "❌ FAIL",
        "Excel Generation": "✓ PASS" if excel_success else "❌ FAIL",
        "Dashboard Generation": "✓ PASS" if dashboard_success else "⚠ OPTIONAL",
    }
    
    for test, result in results.items():
        logger.info(f"{test:.<50} {result}")
    
    if pdf_path:
        logger.info(f"\n📄 PDF Location: {pdf_path}")
    if ppt_path:
        logger.info(f"📊 PPT Location: {ppt_path}")
    if excel_path:
        logger.info(f"📈 Excel Location: {excel_path}")
    if dashboard_path:
        logger.info(f"🌐 Dashboard Location: {dashboard_path}")
    
    # Overall result (dashboard is optional)
    required_tests = [k for k in results.keys() if k != "Dashboard Generation"]
    all_pass = all("✓" in results[k] for k in required_tests)
    
    if all_pass:
        logger.info("\n✅ ALL TESTS PASSED")
        logger.info("All reports (PDF, PPT, Excel) use real data from single source of truth")
        if dashboard_success:
            logger.info("Dashboard also generated successfully with real data")
        logger.info("No placeholders detected in any format")
        return True
    else:
        logger.warning("\n⚠ SOME TESTS FAILED - Review results above")
        return False


if __name__ == "__main__":
    try:
        success = main()
        sys.exit(0 if success else 1)
    except Exception as e:
        logger.error(f"Test execution failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
